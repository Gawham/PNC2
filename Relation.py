import os
import pandas as pd
from google import genai
import boto3
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import PatternFill
import time

# Get the API key from AWS Secrets Manager
def get_secret():
    secret_name = "gemini/api-key"
    region_name = "ap-south-1"

    session = boto3.session.Session()
    client = session.client(
        service_name='secretsmanager',
        region_name=region_name
    )

    try:
        get_secret_value_response = client.get_secret_value(
            SecretId=secret_name
        )
        return get_secret_value_response['SecretString']
    except Exception as e:
        raise e

# Get the API key from Secrets Manager
api_key = get_secret()

client = genai.Client(api_key=api_key)

# Read the CSV file
csv_file_path = 'enriched_pnc.csv'
df = pd.read_csv(csv_file_path)

# Add a new column for relationship
df['Relationship'] = ""

# Process all entries
print("Analyzing name relationships...")
for index, row in df.iterrows():
    if pd.notna(row['OWNER_NAMES']) and pd.notna(row['deceased_name']):
        print(f"Processing record {index+1}/{len(df)}")
        
        prompt = f"""
        Compare these two names and determine the relationship:
        Owner Name: {row['OWNER_NAMES']}
        Deceased Name: {row['deceased_name']}
        
        Respond with exactly one of these options:
        - "owner" if the deceased is one of the owners
        - "family" if the deceased appears to be family to one of the owners
        - "unrelated" if there's no apparent relationship
        - "company" if either name is a company or LLC or Trust
        
        Only respond with one of these exact words, nothing else.
        """
        
        while True:
            try:
                response = client.models.generate_content(
                    model="gemini-1.5-flash",
                    contents=[prompt]
                )
                break
            except Exception as e:
                if "503" in str(e):
                    print("Service unavailable, waiting 5 seconds before retry...")
                    time.sleep(5)
                    continue
                raise e
        
        relationship = response.text.strip().lower()
        print(f"Relationship: {relationship}")
        
        # Update the relationship in the dataframe
        df.at[index, 'Relationship'] = relationship

# Save the updated dataframe to a new file
output_file_path = 'Relationship_Results.xlsx'
df.to_excel(output_file_path, index=False)

# Create workbook and apply formatting
wb = load_workbook(output_file_path)
ws = wb.active

# Create green fill similar to the one in the phone validation sheet (lighter green)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Apply green fill to cells with owner or family relationship
relationship_col_idx = df.columns.get_loc('Relationship') + 1  # +1 because openpyxl is 1-indexed
for row_idx, relationship in enumerate(df['Relationship'], start=2):  # start=2 to skip header
    if relationship in ['owner', 'family']:
        ws.cell(row=row_idx, column=relationship_col_idx).fill = green_fill

wb.save(output_file_path)

print("Relationship analysis complete and saved to Relationship_Results.xlsx")