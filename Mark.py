import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def mark_and_sort_rep_family():
    # Load the workbook
    wb = openpyxl.load_workbook('PNC 3rd May.xlsx', data_only=True)
    sheet = wb.active
    
    # Find the column index for "Is Rep Family?"
    header_row = 1
    rep_family_col = None
    
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=header_row, column=col).value == "Is Rep Family?":
            rep_family_col = col
            break
    
    if rep_family_col is None:
        print("Could not find 'Is Rep Family?' column")
        return

    # Collect all rows (excluding header)
    rows_data = []
    for row in range(2, sheet.max_row + 1):
        row_data = []
        is_rep_family = sheet.cell(row=row, column=rep_family_col).value
        is_true = isinstance(is_rep_family, bool) and is_rep_family
        
        # Store all cell values and formatting in the row
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            row_data.append({
                'value': cell.value,
                'style': cell._style
            })
        rows_data.append((is_true, row_data))
    
    # Sort rows (TRUE values first)
    rows_data.sort(key=lambda x: (not x[0]))  # Sort by is_true in descending order
    
    # Clear the sheet (except header)
    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = None
            
    # Write back the sorted data
    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    true_count = 0
    
    for idx, (is_true, row_data) in enumerate(rows_data, start=2):
        for col_idx, cell_data in enumerate(row_data, start=1):
            cell = sheet.cell(row=idx, column=col_idx)
            cell.value = cell_data['value']
            cell._style = cell_data['style']
            
            if is_true:
                cell.fill = green_fill
                true_count += 1 if col_idx == 1 else 0
    
    print(f"Moved and colored {true_count} rows")
    
    # Save the modified workbook with a new name
    wb.save('PNC 3rd May_sorted.xlsx')
    print("File saved successfully")

if __name__ == "__main__":
    mark_and_sort_rep_family()
